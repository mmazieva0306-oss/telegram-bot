import logging
import os
from datetime import datetime
from flask import Flask, request
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ConversationHandler, CallbackQueryHandler, ContextTypes, filters,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# НАСТРОЙКИ
# ============================================================
BOT_TOKEN = "8716526377:AAHkB-fUW7Mjnixr3JvJVl6tv-DOp70n1I0"
ADMIN_CHAT_ID = "829964557"
EXCEL_FILE = "/tmp/zayavki.xlsx"

REGION, PRODUCT, PRICE, VOLUME, CONTACT, CONFIRM = range(6)

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================
# РЕГИОНЫ (АЛФАВИТНЫЙ ПОРЯДОК)
# ============================================================
REGIONS = [
    ["Алтайский край", "Архангельская область"],
    ["Владимирская область", "Вологодская область"],
    ["Кировская область", "Ленинградская область"],
    ["Пермский край", "Республика Карелия"],
    ["Республика Марий Эл", "✏️ Другой регион"],
]

# ============================================================
# ПРОДУКТЫ
# ============================================================
PRODUCTS = [
    ["🌲 Шишка сосновая", "🍓 Морошка"],
    ["🫐 Черника", "🍊 Облепиха"],
    ["🍓 Земляника", "🍒 Клюква"],
    ["🍓 Клубника", "🍒 Брусника"],
    ["🍓 Малина", "✏️ Другое"],
]

# ============================================================
# EXCEL
# ============================================================
HEADER_COLS = ["№", "Дата", "Регион", "Продукт", "Цена (руб/кг)", "Объём (кг)", "Контакт", "Telegram", "ID"]
COL_WIDTHS = [5, 12, 22, 22, 16, 12, 24, 20, 14]

def _style_header(ws):
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    h_fill = PatternFill("solid", start_color="2E7D32")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    for col, (h, w) in enumerate(zip(HEADER_COLS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = h_font, h_fill, center, thin
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Все заявки"
    _style_header(ws)
    wb.save(EXCEL_FILE)
    logger.info("Создан новый файл Excel")

def save_to_excel(data: dict, user):
    try:
        init_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception as e:
        logger.error(f"Ошибка загрузки Excel: {e}")
        return False
    
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    now = datetime.now()
    
    # Сохраняем в общий лист
    if "Все заявки" not in wb.sheetnames:
        ws_all = wb.create_sheet("Все заявки")
        _style_header(ws_all)
    else:
        ws_all = wb["Все заявки"]
    
    row_num = ws_all.max_row + 1
    fill_all = PatternFill("solid", start_color="F1F8E9" if (row_num - 1) % 2 == 0 else "FFFFFF")
    
    row_data = [
        row_num - 1,
        now.strftime("%d.%m.%Y %H:%M"),
        data.get("region", ""),
        data.get("product", ""),
        data.get("price", ""),
        data.get("volume", ""),
        data.get("contact", ""),
        f"@{user.username}" if user.username else "—",
        str(user.id)
    ]
    
    for col, val in enumerate(row_data, 1):
        cell = ws_all.cell(row=row_num, column=col, value=val)
        cell.font = font
        cell.border = thin
        cell.fill = fill_all
        if col in (1, 2, 5, 6):
            cell.alignment = center
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Сохраняем в лист региона
    region_name = data.get("region", "Без региона")
    safe_name = "".join(c for c in region_name if c not in r'\/:*?"<>|')[:31]
    
    if safe_name not in wb.sheetnames:
        ws_region = wb.create_sheet(title=safe_name)
        _style_header(ws_region)
    else:
        ws_region = wb[safe_name]
    
    row_num_region = ws_region.max_row + 1
    fill_region = PatternFill("solid", start_color="F1F8E9" if (row_num_region - 1) % 2 == 0 else "FFFFFF")
    
    for col, val in enumerate(row_data, 1):
        cell = ws_region.cell(row=row_num_region, column=col, value=val)
        cell.font = font
        cell.border = thin
        cell.fill = fill_region
        if col in (1, 2, 5, 6):
            cell.alignment = center
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    wb.save(EXCEL_FILE)
    logger.info(f"✅ Заявка сохранена в Excel")
    return True

# ============================================================
# ОБРАБОТЧИКИ КОМАНД
# ============================================================
async def send_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if str(update.effective_user.id) != ADMIN_CHAT_ID:
        await update.message.reply_text("⛔ У вас нет прав.")
        return
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("📭 Пока нет ни одной заявки.")
        return
    with open(EXCEL_FILE, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=f"zayavki_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            caption=f"📊 Все заявки — {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        )

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text(
        "👋 Добро пожаловать в бот *Дикоросы России!*\n\n"
        "Оставьте заявку — мы свяжемся с вами.\n\n"
        "━━━━━━━━━━━━━━━\n"
        "📍 *Шаг 1 из 5 — Регион*\n"
        "Выберите ваш регион:",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(REGIONS, one_time_keyboard=True, resize_keyboard=True),
    )
    return REGION

async def get_region(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "✏️ Другой регион":
        await update.message.reply_text("✏️ Напишите ваш регион:", reply_markup=ReplyKeyboardRemove())
        return REGION
    context.user_data["region"] = text
    await update.message.reply_text(
        "━━━━━━━━━━━━━━━\n"
        "🌲 *Шаг 2 из 5 — Продукт*\n"
        "Что хотите сдать?",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(PRODUCTS, one_time_keyboard=True, resize_keyboard=True),
    )
    return PRODUCT

async def get_product(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "✏️ Другое":
        await update.message.reply_text("✏️ Напишите название продукта:", reply_markup=ReplyKeyboardRemove())
        return PRODUCT
    context.user_data["product"] = text
    await update.message.reply_text(
        "━━━━━━━━━━━━━━━\n"
        "💰 *Шаг 3 из 5 — Цена*\n"
        "По какой цене сдаёте? (рублей за 1 кг)\n\nПример: 200",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )
    return PRICE

async def get_price(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["price"] = update.message.text.strip()
    await update.message.reply_text(
        "━━━━━━━━━━━━━━━\n"
        "⚖️ *Шаг 4 из 5 — Объём*\n"
        "Сколько килограмм готовы сдать?\n\nПример: 20",
        parse_mode="Markdown",
    )
    return VOLUME

async def get_volume(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["volume"] = update.message.text.strip()
    await update.message.reply_text(
        "━━━━━━━━━━━━━━━\n"
        "📞 *Шаг 5 из 5 — Контакт*\n"
        "Укажите телефон или Telegram для связи:\n\nПример: +7 900 123-45-67",
        parse_mode="Markdown",
    )
    return CONTACT

async def get_contact(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["contact"] = update.message.text.strip()
    d = context.user_data
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Отправить", callback_data="confirm"),
        InlineKeyboardButton("🔄 Заново", callback_data="restart"),
    ]])
    await update.message.reply_text(
        f"━━━━━━━━━━━━━━━\n"
        f"📋 *Проверьте заявку:*\n\n"
        f"📍 Регион: *{d['region']}*\n"
        f"🌲 Продукт: *{d['product']}*\n"
        f"💰 Цена: *{d['price']} руб/кг*\n"
        f"⚖️ Объём: *{d['volume']} кг*\n"
        f"📞 Контакт: *{d['contact']}*\n\n"
        f"Всё верно?",
        parse_mode="Markdown",
        reply_markup=keyboard,
    )
    return CONFIRM

async def confirm_order(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    
    if query.data == "restart":
        await query.edit_message_text("🔄 Начинаем заново...")
        context.user_data.clear()
        await query.message.reply_text(
            "📍 *Шаг 1 из 5 — Регион*\nВыберите ваш регион:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(REGIONS, one_time_keyboard=True, resize_keyboard=True),
        )
        return REGION
    
    user = update.effective_user
    d = context.user_data
    
    excel_ok = save_to_excel(d, user)
    
    await query.edit_message_text(
        f"✅ *Заявка принята! Спасибо!*\n\n"
        f"📍 {d['region']}\n"
        f"🌲 {d['product']}\n"
        f"💰 {d['price']} руб/кг\n"
        f"⚖️ {d['volume']} кг\n"
        f"📞 {d['contact']}\n\n"
        f"Мы свяжемся с вами. 🙏\n\n"
        f"_/start — новая заявка_",
        parse_mode="Markdown",
    )
    
    admin_text = (
        f"📬 *Новая заявка!*\n"
        f"🕐 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
        f"📍 Регион: *{d['region']}*\n"
        f"🌲 Продукт: *{d['product']}*\n"
        f"💰 Цена: *{d['price']} руб/кг*\n"
        f"⚖️ Объём: *{d['volume']} кг*\n"
        f"📞 Контакт: *{d['contact']}*\n\n"
        f"👤 {user.full_name} (@{user.username or '—'})\n"
        f"🆔 `{user.id}`\n\n"
        f"{'✅ Сохранено в Excel' if excel_ok else '❌ Ошибка сохранения Excel'}"
    )
    try:
        await context.bot.send_message(chat_id=int(ADMIN_CHAT_ID), text=admin_text, parse_mode="Markdown")
    except Exception as e:
        logger.error(f"Ошибка отправки админу: {e}")
    
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("❌ Отменено. /start", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# ============================================================
# FLASK ДЛЯ ВЕБХУКА
# ============================================================
flask_app = Flask(__name__)
telegram_app = None

@flask_app.route(f'/{BOT_TOKEN}', methods=['POST'])
async def webhook():
    try:
        update = Update.de_json(request.get_json(force=True), telegram_app.bot)
        await telegram_app.process_update(update)
        return 'OK', 200
    except Exception as e:
        logger.error(f"Ошибка вебхука: {e}")
        return 'Error', 500

@flask_app.route('/')
def health_check():
    return 'Бот работает', 200

def setup_webhook():
    render_url = os.environ.get('RENDER_EXTERNAL_URL')
    if not render_url:
        logger.warning("RENDER_EXTERNAL_URL не найден")
        return False
    
    webhook_url = f"{render_url}/{BOT_TOKEN}"
    telegram_app.bot.set_webhook(url=webhook_url, drop_pending_updates=True)
    logger.info(f"✅ Вебхук установлен: {webhook_url}")
    return True

# ============================================================
# ЗАПУСК
# ============================================================
def main():
    global telegram_app
    
    print("🚀 Запуск бота...")
    init_excel()
    
    telegram_app = Application.builder().token(BOT_TOKEN).build()
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            REGION: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_region)],
            PRODUCT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_product)],
            PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_price)],
            VOLUME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_volume)],
            CONTACT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_contact)],
            CONFIRM: [CallbackQueryHandler(confirm_order)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    telegram_app.add_handler(conv_handler)
    telegram_app.add_handler(CommandHandler("excel", send_excel))
    
    if setup_webhook():
        port = int(os.environ.get('PORT', 8080))
        logger.info(f"🚀 Запуск сервера на порту {port}")
        # Запускаем Flask с Gunicorn (не используем app.run())
        flask_app.run(host='0.0.0.0', port=port)
    else:
        logger.warning("Запуск в режиме polling")
        telegram_app.run_polling()

# ============================================================
# ТОЧКА ВХОДА ДЛЯ GUNICORN
# ============================================================
if __name__ == "__main__":
    main()

# Gunicorn будет использовать flask_app напрямую
